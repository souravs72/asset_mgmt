"""Convert legacy client XLSX exports into ERPNext import CSV files.

Run with bench, for example:
bench --site SITE execute asset_mgmt.import.transform_legacy.run --kwargs '{"output_dir": "/tmp/asset-import", "limit": 100}'
"""

from __future__ import annotations

import csv
import os
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any


DEFAULT_SOURCE_DIR = Path("/home/ascra/Downloads/asset-drive")
SOURCE_DIR_ENV = "ASSET_MGMT_LEGACY_SOURCE_DIR"
EXPORT_SHEET = "Export Worksheet"

ASSET_FILE = "Asset_14052026.xlsx"
CATEGORY_FILE = "Category_list.xlsx"
LOCATION_FILE = "Location_list.xlsx"
COST_CENTER_FILE = "CostCenter_list.xlsx"
SUPPLIER_FILE = "Supp_list.xlsx"
ENRICHMENT_FILE = "FORMAT - Assets Counting.xlsx"

ASSET_TEMPLATE = Path(__file__).resolve().parents[1] / "fixtures" / "import_templates" / "05_asset.csv"

DEFAULT_LEGACY_PURCHASE_DATE = "2000-01-01"

ASSET_EXTRA_COLUMNS = [
	"cost_center",
	"supplier",
	"serial_number",
	"legacy_asset_code",
	"legacy_group",
	"legacy_category",
	"legacy_subcategory",
	"asset_condition",
	"legacy_import_notes",
]


@dataclass(frozen=True)
class CompanySettings:
	company_name: str = "Asset Management"
	company_abbr: str = "AM"


def run(output_dir, limit=None):
	"""Bench execute entrypoint for generating all legacy import CSVs."""
	masters = transform_masters(output_dir)
	asset = transform_assets(output_dir, limit=limit)

	return {
		"output_dir": str(Path(output_dir).expanduser()),
		"masters": masters,
		"asset": asset,
	}


def transform_masters(output_dir):
	"""Generate master CSV files for Cost Center, Location, Asset Category, Item and Supplier."""
	settings = _get_company_settings()
	output_path = _ensure_output_dir(output_dir)

	categories: dict[str, None] = {}
	items: dict[tuple[str, str], str] = {}
	locations: dict[str, None] = {}
	cost_centers: dict[str, None] = {}
	suppliers: dict[str, None] = {}

	for row in _read_rows(CATEGORY_FILE):
		group = _clean_text(row.get("TDFA_GROUP_NAME_P"))
		subcategory = _clean_text(row.get("TDFA_SCATEGORY_NAME_P"))
		if group:
			categories.setdefault(group, None)
		if subcategory:
			items.setdefault((subcategory, group), _slug(subcategory))

	for row in _read_rows(LOCATION_FILE):
		location = _clean_text(row.get("TDFA_LOCATION_NAME_P"))
		if location:
			locations.setdefault(location, None)

	for row in _read_rows(COST_CENTER_FILE):
		cost_center = _clean_text(row.get("TDFA_CC_NAME_P"))
		if cost_center:
			cost_centers.setdefault(cost_center, None)

	for row in _read_rows(SUPPLIER_FILE):
		supplier = _clean_text(row.get("TD_BEN_DESC_P"))
		if supplier:
			suppliers.setdefault(supplier, None)

	for row in _read_rows(ASSET_FILE):
		group = _clean_text(row.get("TDFA_GROUP_NAME_P"))
		subcategory = _clean_text(row.get("TDFA_SCATEGORY_NAME_P"))
		location = _clean_text(row.get("TDFA_LOCATION_NAME_P"))
		cost_center = _clean_text(row.get("TDFA_CC_NAME_P"))
		supplier = _clean_text(row.get("TD_BEN_DESC_P"))

		if group:
			categories.setdefault(group, None)
		if subcategory:
			items.setdefault((subcategory, group), _slug(subcategory))
		if location:
			locations.setdefault(location, None)
		if cost_center:
			cost_centers.setdefault(cost_center, None)
		if supplier:
			suppliers.setdefault(supplier, None)

	_resolve_item_code_collisions(items)

	files = {
		"cost_center": _write_csv(
			output_path / "cost_center.csv",
			["cost_center_name", "parent_cost_center", "company", "is_group"],
			(
				{
					"cost_center_name": name,
					"parent_cost_center": _company_cost_center(settings),
					"company": settings.company_name,
					"is_group": 0,
				}
				for name in cost_centers
			),
		),
		"location": _write_csv(
			output_path / "location.csv",
			["location_name", "parent_location", "is_group"],
			({"location_name": name, "parent_location": "", "is_group": 0} for name in locations),
		),
		"asset_category": _write_csv(
			output_path / "asset_category.csv",
			[
				"asset_category_name",
				"fixed_asset_account",
				"accumulated_depreciation_account",
				"depreciation_expense_account",
				"depreciation_method",
				"total_number_of_depreciations",
				"frequency_of_depreciation",
			],
			(
				{
					"asset_category_name": name,
					"fixed_asset_account": _fixed_asset_account(name, settings),
					"accumulated_depreciation_account": _account("Accumulated Depreciations", settings),
					"depreciation_expense_account": _account("Depreciation", settings),
					"depreciation_method": "Straight Line",
					"total_number_of_depreciations": 60,
					"frequency_of_depreciation": 1,
				}
				for name in categories
			),
		),
		"item": _write_csv(
			output_path / "item.csv",
			["item_code", "item_name", "item_group", "stock_uom", "is_fixed_asset", "is_stock_item", "asset_category"],
			(
				{
					"item_code": item_code,
					"item_name": subcategory,
					"item_group": "All Item Groups",
					"stock_uom": "Nos",
					"is_fixed_asset": 1,
					"is_stock_item": 0,
					"asset_category": group,
				}
				for (subcategory, group), item_code in items.items()
			),
		),
		"supplier": _write_csv(
			output_path / "supplier.csv",
			["supplier_name", "supplier_group", "supplier_type"],
			(
				{
					"supplier_name": name,
					"supplier_group": "All Supplier Groups",
					"supplier_type": "Company",
				}
				for name in suppliers
			),
		),
	}

	return files


def transform_assets(output_dir, limit=None):
	"""Generate the Asset import CSV from the legacy asset workbook."""
	settings = _get_company_settings()
	output_path = _ensure_output_dir(output_dir)
	columns = _asset_columns()
	enrichment = _load_enrichment()
	row_limit = _normalise_limit(limit)
	seen_asset_codes: set[str] = set()
	duplicate_codes: list[str] = []

	def asset_rows():
		for index, row in enumerate(_read_rows(ASSET_FILE), start=1):
			if row_limit is not None and index > row_limit:
				break

			asset_code = _clean_text(row.get("TDFA_ASSET_CODE"))
			if asset_code:
				if asset_code in seen_asset_codes:
					duplicate_codes.append(asset_code)
					continue
				seen_asset_codes.add(asset_code)
			extra = enrichment.get(asset_code, {})
			group = _clean_text(row.get("TDFA_GROUP_NAME_P"))
			category = _clean_text(row.get("TDFA_CATEGORY_NAME_P"))
			subcategory = _clean_text(row.get("TDFA_SCATEGORY_NAME_P"))
			asset_name = _clean_text(row.get("TDFA_ASSET_NAME_P")) or _clean_text(extra.get("Asset Name")) or asset_code
			operational_status, asset_condition = _map_status(
				row.get("TDFA_ASSET_STATUS"),
				row.get("TDFA_ASSET_ACTIVE"),
				extra.get("Status"),
			)
			depreciation_percent = _clean_text(row.get("TDFA_ASSET_DEPRN_PERC")) or _clean_text(
				extra.get("Depreciation %")
			)
			gross_purchase_amount = _decimal_string(row.get("TDFA_ASSET_CAPITAL_COST")) or _decimal_string(
				extra.get("Capital Cost")
			)
			depreciation_start_date = _date_string(extra.get("Depreciation Start Date")) or DEFAULT_LEGACY_PURCHASE_DATE

			yield {
				"asset_name": asset_name,
				"item_code": _slug(subcategory or category or group or asset_code),
				"asset_category": group,
				"company": settings.company_name,
				"location": _clean_text(row.get("TDFA_LOCATION_NAME_P")) or _clean_text(extra.get("Location Name")),
				"custodian": "",
				"purchase_date": depreciation_start_date,
				"available_for_use_date": depreciation_start_date,
				"gross_purchase_amount": gross_purchase_amount,
				"is_existing_asset": 1,
				"calculate_depreciation": 1 if depreciation_percent else 0,
				"opening_accumulated_depreciation": _decimal_string(extra.get("Depreciation Amount")),
				"asset_tag": asset_code,
				"asset_tag_type": "Barcode",
				"operational_status": operational_status,
				"cost_center": _clean_text(row.get("TDFA_CC_NAME_P")) or _clean_text(extra.get("Cost Center Name")),
				"supplier": _clean_text(row.get("TD_BEN_DESC_P")) or _clean_text(extra.get("Supplier Name")),
				"serial_number": _clean_text(row.get("TDFA_ASSET_SRL_NO")),
				"legacy_asset_code": asset_code,
				"legacy_group": group,
				"legacy_category": category,
				"legacy_subcategory": subcategory,
				"asset_condition": asset_condition,
				"legacy_import_notes": _legacy_import_notes(depreciation_percent, extra),
			}

	output_file = _write_csv(output_path / "asset.csv", columns, asset_rows())
	return {
		"path": output_file,
		"rows_written": len(seen_asset_codes),
		"duplicate_codes_skipped": len(duplicate_codes),
		"duplicate_examples": duplicate_codes[:10],
	}


def _source_dir() -> Path:
	return Path(os.environ.get(SOURCE_DIR_ENV, DEFAULT_SOURCE_DIR)).expanduser()


def _ensure_output_dir(output_dir) -> Path:
	output_path = Path(output_dir).expanduser()
	output_path.mkdir(parents=True, exist_ok=True)
	return output_path


def _get_company_settings() -> CompanySettings:
	try:
		from asset_mgmt.settings import get_settings

		settings = get_settings()
		return CompanySettings(
			company_name=_setting(settings, "company_name", "Asset Management"),
			company_abbr=_setting(settings, "company_abbr", "AM"),
		)
	except Exception:
		return CompanySettings()


def _setting(settings: Any, key: str, default: str) -> str:
	if isinstance(settings, dict):
		return _clean_text(settings.get(key)) or default
	return _clean_text(getattr(settings, key, None)) or default


def _read_rows(file_name: str):
	path = _source_dir() / file_name
	if not path.exists():
		raise FileNotFoundError(f"Legacy source workbook not found: {path}")

	try:
		from openpyxl import load_workbook
	except ImportError:
		raise ImportError("openpyxl is required to read the legacy XLSX files. Install it in the bench environment.")

	workbook = load_workbook(path, read_only=True, data_only=True)
	try:
		worksheet = workbook[EXPORT_SHEET] if EXPORT_SHEET in workbook.sheetnames else workbook[workbook.sheetnames[0]]
		rows = worksheet.iter_rows(values_only=True)
		headers = [_clean_header(value) for value in next(rows)]

		for values in rows:
			if not any(_clean_text(value) for value in values):
				continue
			yield {header: value for header, value in zip(headers, values, strict=False) if header}
	finally:
		workbook.close()


def _load_enrichment() -> dict[str, dict[str, Any]]:
	path = _source_dir() / ENRICHMENT_FILE
	if not path.exists():
		return {}

	enrichment = {}
	for row in _read_rows(ENRICHMENT_FILE):
		asset_code = _clean_text(row.get("Asset Code"))
		if asset_code:
			enrichment[asset_code] = row
	return enrichment


def _asset_columns() -> list[str]:
	with ASSET_TEMPLATE.open(newline="", encoding="utf-8") as handle:
		reader = csv.reader(handle)
		columns = next(reader)

	for column in ASSET_EXTRA_COLUMNS:
		if column not in columns:
			columns.append(column)

	return columns


def _write_csv(path: Path, fieldnames: list[str], rows) -> str:
	path.parent.mkdir(parents=True, exist_ok=True)
	with path.open("w", newline="", encoding="utf-8") as handle:
		writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
		writer.writeheader()
		for row in rows:
			writer.writerow({fieldname: _csv_value(row.get(fieldname, "")) for fieldname in fieldnames})
	return str(path)


def _clean_header(value: Any) -> str:
	return _clean_text(value).replace("\ufeff", "")


def _clean_text(value: Any) -> str:
	if value is None:
		return ""
	if isinstance(value, datetime):
		return value.date().isoformat()
	if isinstance(value, date):
		return value.isoformat()
	text = str(value).strip()
	return re.sub(r"\s+", " ", text)


def _csv_value(value: Any) -> str:
	if isinstance(value, int | float | Decimal):
		return _decimal_string(value)
	return _clean_text(value)


def _decimal_string(value: Any) -> str:
	text = _clean_text(value).replace(",", "")
	if not text:
		return ""
	try:
		decimal = Decimal(text)
	except InvalidOperation:
		return text
	return format(decimal.normalize(), "f")


def _date_string(value: Any) -> str:
	if isinstance(value, datetime):
		return value.date().isoformat()
	if isinstance(value, date):
		return value.isoformat()

	text = _clean_text(value)
	if not text:
		return ""
	for date_format in ("%d/%m/%Y", "%d-%b-%y", "\\%d-%b-%y", "%Y-%m-%d"):
		try:
			return datetime.strptime(text, date_format).date().isoformat()
		except ValueError:
			pass

	try:
		serial = int(Decimal(text))
	except (InvalidOperation, ValueError):
		return text
	return (date(1899, 12, 30) + timedelta(days=serial)).isoformat()


def _slug(value: str, fallback: str = "UNKNOWN") -> str:
	text = _clean_text(value).upper()
	slug = re.sub(r"[^A-Z0-9]+", "-", text).strip("-")
	return (slug or fallback)[:90]


def _resolve_item_code_collisions(items: dict[tuple[str, str], str]) -> None:
	seen: dict[str, tuple[str, str]] = {}
	for key, item_code in list(items.items()):
		if item_code not in seen:
			seen[item_code] = key
			continue

		group = key[1]
		group_suffix = _slug(group)[:20]
		new_code = f"{item_code[:68]}-{group_suffix}" if group_suffix else item_code
		counter = 2
		while new_code in seen:
			new_code = f"{item_code[:84]}-{counter}"
			counter += 1
		items[key] = new_code
		seen[new_code] = key


def _map_status(status: Any, active: Any, enriched_status: Any = None) -> tuple[str, str]:
	status_code = _clean_text(status).upper()
	active_code = _clean_text(active).upper()
	enriched = _clean_text(enriched_status)

	if status_code == "I" and active_code == "A":
		return "Active", "In Service"
	if status_code == "I":
		return "Active", enriched or "In Service"
	if status_code == "SO":
		return "Retired", "Sold"
	if status_code == "SC":
		return "Retired", "Scrapped"
	if status_code == "D" or active_code == "D":
		return "Under Maintenance", "Damaged"
	if enriched:
		return "Active", enriched
	if active_code in {"A", "Y"}:
		return "Active", "In Service"
	return "Active", "In Service"


def _legacy_import_notes(depreciation_percent: str, enrichment: dict[str, Any]) -> str:
	parts = []
	if depreciation_percent:
		parts.append(f"Legacy depreciation %: {depreciation_percent}")

	old_code = _clean_text(enrichment.get("Asset Old Code"))
	if old_code:
		parts.append(f"Legacy old code: {old_code}")

	book_value = _decimal_string(enrichment.get("Book Value"))
	if book_value:
		parts.append(f"Legacy book value: {book_value}")

	return "; ".join(parts)


def _normalise_limit(limit: Any) -> int | None:
	if limit in (None, ""):
		return None
	limit_value = int(limit)
	if limit_value < 0:
		raise ValueError("limit must be greater than or equal to zero")
	if limit_value == 0:
		return None
	return limit_value


def _account(name: str, settings: CompanySettings) -> str:
	return f"{name} - {settings.company_abbr}"


def _company_cost_center(settings: CompanySettings) -> str:
	return f"{settings.company_name} - {settings.company_abbr}"


def _fixed_asset_account(category_name: str, settings: CompanySettings) -> str:
	name = category_name.upper()
	if "FURNIT" in name or "FIXTURE" in name:
		return _account("Furnitures and Fixtures", settings)
	if "HARDWARE" in name or "COMPUTER" in name or name.startswith("IT "):
		return _account("Electronic Equipments", settings)
	return _account("Capital Equipments", settings)
